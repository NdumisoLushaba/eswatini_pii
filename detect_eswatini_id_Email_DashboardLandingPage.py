import os
import re
import pandas as pd
import smtplib
from datetime import datetime
import hashlib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import pytesseract
from pdf2image import convert_from_path
import pdfplumber
import openpyxl
from PIL import Image
import logging
import warnings
import tkinter as tk
from tkinter import messagebox
import threading
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from transformers import pipeline
import fasttext
import requests
from langdetect import detect, DetectorFactory

# Initialize deterministic language detection
DetectorFactory.seed = 0

# Suppress non-critical warnings
logging.getLogger("pdfminer").setLevel(logging.ERROR)
warnings.filterwarnings("ignore")

# Tesseract path
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Folders
FOLDER_PATH       = r"C:\Users\nlushaba\PII_Project\scans"         # Manual / batch scans
LOG_FOLDER        = r"C:\Users\nlushaba\PII_Project\logs"          # Where Excel logs go
USB_MONITOR_PATH  = r"C:\Users\nlushaba\PII_Project\USB_DRIVE"     # Folder to monitor for moved files

os.makedirs(LOG_FOLDER, exist_ok=True)

# Email settings
SENDER_EMAIL   = "ndumisolushaba0@gmail.com"
PASSWORD       = "bugmmsqggpijxmtd"
RECIPIENT_EMAIL= "ndumisolushaba0@gmail.com"
CC_EMAILS      = []

# Load FastText language detection model
LANG_MODEL_PATH = "lid.176.bin"
if not os.path.exists(LANG_MODEL_PATH):
    url = "https://dl.fbaipublicfiles.com/fasttext/supervised-models/lid.176.bin"
    r = requests.get(url, allow_redirects=True)
    with open(LANG_MODEL_PATH, 'wb') as f:
        f.write(r.content)
lang_model = fasttext.load_model(LANG_MODEL_PATH)

# Initialize NER pipeline
ner_pipeline = pipeline(
    "ner", 
    model="Davlan/bert-base-multilingual-cased-ner-hrl",  # Multilingual model
    aggregation_strategy="max",
    device=-1  # Use CPU (-1) or GPU (0)
)

# Base PII Patterns (language-independent)
BASE_PATTERNS = {
    "National ID": r"\b\d{6}(?:6100|1100|2100|7100)\d{3}\b",
    "Phone Number": r"\b(?:\+268)?(?:7[89]\d{6}|[56]\d{7}|2\d{7})\b",
    "Passport Number": r"\b[A-Z]{1,2}\d{6,8}\b",
    "TIN": r"\b\d{9}\b",
    "Driver’s License": r"\bD\d{7,9}\b",
    "Bank Account": r"\b(?:62|02|91|77|10|99|63)\d{8,10}\b",
    "Credit Card": r"\b(?:\d{4}[ -]?){3}\d{4}\b",
    "Health Record": r"\b(?:MED|PAT)\d{6,8}\b",
    "Student ID": r"\b\d{8,10}\b"
}

# Language-specific regex enhancements
LANGUAGE_REGEX_MAP = {
    "en": {
        "Email": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
        "Name Prefix": r"\b(?:Mr|Mrs|Ms|Dr|Prof)\.?\b"
    },
    "fr": {
        "Email": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
        "Name Prefix": r"\b(?:M|Mme|Mlle|Dr|Prof)\.?\b"
    },
    "es": {
        "Email": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
        "Name Prefix": r"\b(?:Sr|Sra|Srta|Dr|Prof)\.?\b"
    },
    "pt": {
        "Email": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
        "Name Prefix": r"\b(?:Sr|Sra|Srta|Dr|Prof)\.?\b"
    },
    "zu": {
        "Phone Number": r"\b(?:\+27|0)[1-8]\d{8}\b"
    }
}

def get_log_path():
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return os.path.join(LOG_FOLDER, f"detected_pii_{ts}.xlsx")

def detect_language(text):
    """Detect language using multiple methods for reliability"""
    if len(text) < 50:
        return "unknown"
    
    try:
        # Method 1: FastText (most reliable for short texts)
        lang_pred = lang_model.predict(text.replace("\n", " "))[0][0].split("__")[-1]
        if lang_pred in LANGUAGE_REGEX_MAP:
            return lang_pred
    except:
        pass
    
    try:
        # Method 2: langdetect
        return detect(text)
    except:
        pass
    
    try:
        # Method 3: langid
        return langid.classify(text)[0]
    except:
        return "unknown"

def get_patterns_for_text(text):
    """Get combined patterns based on detected language"""
    lang = detect_language(text)
    patterns = BASE_PATTERNS.copy()
    
    # Add language-specific patterns if available
    if lang in LANGUAGE_REGEX_MAP:
        for key, pattern in LANGUAGE_REGEX_MAP[lang].items():
            patterns[key] = pattern
    
    return patterns

# Extraction functions with OCR support ------------------------------
def extract_text_from_txt(path):
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    except:
        return ""

def extract_text_from_pdf(path):
    text = ""
    try:
        with pdfplumber.open(path) as pdf:
            for p in pdf.pages:
                text += (p.extract_text() or "") + "\n"
    except:
        pass
    
    # Fallback to OCR if text extraction fails
    if len(text) < 50:
        try:
            images = convert_from_path(path, dpi=300)
            for img in images:
                text += pytesseract.image_to_string(img, lang='eng+fra+spa+por') + "\n"
        except Exception as e:
            print(f"PDF OCR failed: {e}")
    return text

def extract_text_from_doc(path):
    try:
        import docx2txt
        return docx2txt.process(path) or ""
    except:
        return ""

def extract_text_from_xlsx(path):
    content = ""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        content += f"{cell} "
    except:
        pass
    return content

def extract_text_from_image(path):
    try:
        img = Image.open(path)
        # Detect language for OCR optimization
        temp_text = pytesseract.image_to_string(img, lang='eng')
        lang = detect_language(temp_text)
        
        # Use appropriate language packs
        langs = 'eng'
        if lang == 'fr': langs = 'fra'
        elif lang == 'es': langs = 'spa'
        elif lang == 'pt': langs = 'por'
        
        return pytesseract.image_to_string(img, lang=langs)
    except Exception as e:
        print(f"Image OCR failed: {e}")
        return ""

def extract_text_from_any_file(path):
    """
    Attempts PDF → DOC → XLSX → image → TXT extraction in order.
    Returns the first non-empty result.
    """
    ext = path.lower().split('.')[-1]
    
    handlers = {
        'pdf': extract_text_from_pdf,
        'docx': extract_text_from_doc,
        'xlsx': extract_text_from_xlsx,
        'txt': extract_text_from_txt,
        'png': extract_text_from_image,
        'jpg': extract_text_from_image,
        'jpeg': extract_text_from_image,
        'bmp': extract_text_from_image,
        'tiff': extract_text_from_image
    }
    
    for extension, handler in handlers.items():
        if ext == extension:
            return handler(path)
    
    # Fallback for unknown extensions
    for handler in (
        extract_text_from_pdf,
        extract_text_from_doc,
        extract_text_from_xlsx,
        extract_text_from_txt,
        extract_text_from_image
    ):
        try:
            text = handler(path)
            if text and text.strip():
                return text
        except Exception as e:
            print(f"Extraction failed for {path}: {e}")
    
    return ""

# Logging & Email -----------------------------------------------
def log_detected_pii(df, pii_type, value, filename, line_no, confidence=1.0):
    masked = value[:2] + '*'*(len(value)-4) + value[-2:] if len(value)>4 else '*'*len(value)
    entry = {
        "Timestamp": pd.Timestamp.now(),
        "PII Type": pii_type,
        "Masked PII": masked,
        "SHA256 Hash": hashlib.sha256(value.encode()).hexdigest(),
        "File Name": filename,
        "Line Number": line_no,
        "Confidence": confidence
    }
    return pd.concat([df, pd.DataFrame([entry])], ignore_index=True)

def send_email_alert(subject, body, log_df):
    path = get_log_path()
    log_df.to_excel(path, index=False)
    msg = MIMEMultipart()
    msg['From'] = SENDER_EMAIL
    msg['To']   = RECIPIENT_EMAIL
    msg['Subject'] = subject
    if CC_EMAILS:
        msg['Cc'] = ','.join(CC_EMAILS)
    msg.attach(MIMEText(body, 'plain'))
    with open(path, 'rb') as f:
        part = MIMEApplication(f.read(), Name=os.path.basename(path))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(path)}"'
        msg.attach(part)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(SENDER_EMAIL, PASSWORD)
            server.sendmail(SENDER_EMAIL, [RECIPIENT_EMAIL]+CC_EMAILS, msg.as_string())
        print("✅ Email sent:", path)
    except Exception as e:
        print("📨 Email failed:", e)

def show_popup(title, message):
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(title, message)
    root.destroy()

# Folder-scan function (updated) -------------------------------
def scan_folder(queue, folder_path):
    df = pd.DataFrame(columns=[
        "Timestamp","PII Type","Masked PII","SHA256 Hash","File Name","Line Number","Confidence"
    ])

    for fname in os.listdir(folder_path):
        full = os.path.join(folder_path, fname)
        if not os.path.isfile(full):
            continue

        text = extract_text_from_any_file(full)
        if queue:
            queue.put({"type":"file_scanned","file_name":fname})

        if not text.strip():
            print(f"Skipping (no text): {fname}")
            continue

        print(f"\n--- Scanning {fname} ---")
        detected = False
        
        # Get language-specific patterns
        patterns = get_patterns_for_text(text)
        print(f"Detected language: {detect_language(text)}")
        
        # Regex-based detection
        for pii, pat in patterns.items():
            for idx, m in enumerate(re.findall(pat, text, flags=re.IGNORECASE), start=1):
                detected = True
                print(f"  ✅ [Regex] {pii}: {m}")
                df = log_detected_pii(df, pii, m, fname, idx)
                if queue:
                    queue.put({
                        "type":"pii_detected",
                        "pii_type":pii,
                        "file_name":fname,
                        "line":idx
                    })

        # NER-based name detection
        try:
            # Process in chunks for large documents
            chunk_size = 1000
            chunks = [text[i:i+chunk_size] for i in range(0, len(text), chunk_size)]
            
            for chunk_idx, chunk in enumerate(chunks):
                entities = ner_pipeline(chunk)
                
                for entity in entities:
                    if entity['entity_group'] == 'PER' and entity['score'] >= 0.85:
                        detected = True
                        line_no = chunk_idx * chunk_size + entity['start']  # Approximate line number
                        print(f"  ✅ [NER] Personal Name: {entity['word']} (confidence: {entity['score']:.2f})")
                        df = log_detected_pii(
                            df, 
                            "Personal Name", 
                            entity['word'], 
                            fname, 
                            line_no,
                            entity['score']
                        )
                        if queue:
                            queue.put({
                                "type":"pii_detected",
                                "pii_type":"Personal Name",
                                "file_name":fname,
                                "line":line_no
                            })
        except Exception as e:
            print(f"NER processing failed: {e}")

        if not detected:
            print(f"  ❌ No PII in {fname}")

    if not df.empty:
        summary = df.to_string(index=False)
        show_popup("Scan Complete", "PII found—sending email alert now…")
        send_email_alert("✅ PII Detection Alert", f"Dear Team,\n\n{summary}\n\nRegards", df)
    else:
        print("✅ No PII found.")

# USB/folder monitor (updated) ----------------------------------
def try_extract_text_with_retries(path, retries=5, delay=1):
    for i in range(retries):
        text = extract_text_from_any_file(path)
        if text.strip():
            return text
        time.sleep(delay)
    return ""

class PIIFileHandler(FileSystemEventHandler):
    def __init__(self, base_patterns, lang_model):
        super().__init__()
        self.base_patterns = base_patterns
        self.lang_model = lang_model

    def get_patterns_for_text(self, text):
        lang = detect_language(text)
        patterns = self.base_patterns.copy()
        
        if lang in LANGUAGE_REGEX_MAP:
            for key, pattern in LANGUAGE_REGEX_MAP[lang].items():
                patterns[key] = pattern
                
        return patterns

    def detect_pii(self, text, filename):
        # Check regex patterns
        patterns = self.get_patterns_for_text(text)
        for pat in patterns.values():
            if re.search(pat, text, flags=re.IGNORECASE):
                return True
        
        # Check NER for names
        try:
            entities = ner_pipeline(text[:2000])  # Only check first 2000 chars for performance
            for entity in entities:
                if entity['entity_group'] == 'PER' and entity['score'] >= 0.85:
                    return True
        except:
            pass
            
        return False

    def on_created(self, event):
        if event.is_directory:
            return
        path = event.src_path
        print("Detected new file:", path)
        text = try_extract_text_with_retries(path)
        if not text.strip():
            print("  → No text extracted.")
            return

        if self.detect_pii(text, os.path.basename(path)):
            show_popup("PII Detected",
                       f"'{os.path.basename(path)}' contains PII and cannot stay here!")
            try:
                os.remove(path)
                print("  → Deleted:", path)
            except Exception as e:
                print("  → Delete failed:", e)

def start_monitor(folder, base_patterns, lang_model):
    handler  = PIIFileHandler(base_patterns, lang_model)
    observer = Observer()
    observer.schedule(handler, folder, recursive=False)
    observer.start()
    print("Monitoring folder:", folder)
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

# Launch monitor thread and run initial scan --------------------
monitor_thread = threading.Thread(
    target=start_monitor,
    args=(USB_MONITOR_PATH, BASE_PATTERNS, lang_model),
    daemon=True
)
monitor_thread.start()

if __name__ == "__main__":
    # initial batch scan
    scan_folder(None, FOLDER_PATH)

    print("Monitoring in background. Ctrl+C to exit.")
    try:
        while True:
            time.sleep(10)
    except KeyboardInterrupt:
        print("Exiting.")