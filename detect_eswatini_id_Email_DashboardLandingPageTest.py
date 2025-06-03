import os
import re
import pandas as pd
import smtplib
from datetime import datetime
import hashlib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import docx2txt
import pytesseract
from pdf2image import convert_from_path
import pdfplumber
import openpyxl
from PIL import Image
import logging
import warnings
import tkinter as tk
from tkinter import messagebox
import threading
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from google.cloud import dlp
from google.cloud.dlp_v2 import types
import base64

# Suppress warnings
logging.getLogger("pdfminer").setLevel(logging.ERROR)
warnings.filterwarnings("ignore")

# Tesseract path
pytesseract.pytesseract.tesseract_cmd = r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe"

# Folder paths
FOLDER_PATH = r"C:\\Users\\nlushaba\\PII_Project\\scans"
LOG_FOLDER = r"C:\\Users\\nlushaba\\PII_Project\\logs"
USB_MONITOR_PATH = r"C:\\Users\\nlushaba\\PII_Project\\USB_DRIVE"
os.makedirs(LOG_FOLDER, exist_ok=True)

# Email settings
SENDER_EMAIL = "ndumisolushaba0@gmail.com"
PASSWORD = "bugmmsqggpijxmtd"
RECIPIENT_EMAIL = "ndumisolushaba0@gmail.com"
CC_EMAILS = []

# Regex-based patterns
patterns = {
    "National ID": r"\b\d{6}(?:6100|1100|2100|7100)\d{3}\b",
    "Phone Number": r"\b(?:\+268)?(?:7[89]\d{6}|[56]\d{7}|2\d{7})\b",
    "Passport Number": r"\b[A-Z]{1,2}\d{6,8}\b",
    "TIN": r"\b\d{9}\b",
    "Driver’s License": r"\bD\d{7,9}\b",
    "Bank Account": r"\b(?:62|02|91|77|10|99|63)\d{8,10}\b",
    "Credit Card": r"\b(?:\d{4}[ -]?){3}\d{4}\b",
    "Health Record": r"\b(?:MED|PAT)\d{6,8}\b",
    "Student ID": r"\b\d{8,10}\b"
}


def get_log_path():
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return os.path.join(LOG_FOLDER, f"detected_pii_{ts}.xlsx")


# File extraction methods

def extract_text_from_any_file(path):
    for func in (
        extract_text_from_pdf,
        extract_text_from_doc,
        extract_text_from_xlsx,
        extract_text_from_png,
        extract_text_from_txt
    ):
        try:
            text = func(path)
            if text and text.strip():
                return text
        except:
            pass
    return ""


def extract_text_from_txt(path):
    try:
        return open(path, 'r', encoding='utf-8').read()
    except:
        return ""


def extract_text_from_pdf(path):
    text = ""
    try:
        with pdfplumber.open(path) as pdf:
            for p in pdf.pages:
                text += (p.extract_text() or "") + "\n"
    except:
        pass
    if not text.strip():
        try:
            for img in convert_from_path(path):
                text += pytesseract.image_to_string(img)
        except:
            pass
    return text


def extract_text_from_doc(path):
    try:
        return docx2txt.process(path) or ""
    except:
        return ""


def extract_text_from_xlsx(path):
    content = ""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        content += f"{cell} "
    except:
        pass
    return content


def extract_text_from_png(path):
    try:
        return pytesseract.image_to_string(Image.open(path))
    except:
        return ""


# Logging and email

def log_detected_pii(df, pii_type, value, filename, line_no):
    masked = value[:2] + '*'*(len(value)-4) + value[-2:] if len(value) > 4 else '*'*len(value)
    entry = {
        "Timestamp": pd.Timestamp.now(),
        "PII Type": pii_type,
        "Masked PII": masked,
        "SHA256 Hash": hashlib.sha256(value.encode()).hexdigest(),
        "File Name": filename,
        "Line Number": line_no
    }
    return pd.concat([df, pd.DataFrame([entry])], ignore_index=True)


def send_email_alert(subject, body, log_df):
    path = get_log_path()
    log_df.to_excel(path, index=False)
    msg = MIMEMultipart()
    msg['From'] = SENDER_EMAIL
    msg['To'] = RECIPIENT_EMAIL
    msg['Subject'] = subject
    if CC_EMAILS:
        msg['Cc'] = ','.join(CC_EMAILS)
    msg.attach(MIMEText(body, 'plain'))
    with open(path, 'rb') as f:
        part = MIMEApplication(f.read(), Name=os.path.basename(path))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(path)}"'
        msg.attach(part)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(SENDER_EMAIL, PASSWORD)
            server.sendmail(SENDER_EMAIL, [RECIPIENT_EMAIL]+CC_EMAILS, msg.as_string())
        print("✅ Email sent:", path)
    except Exception as e:
        print("❌ Email failed:", e)


def show_popup(title, message):
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(title, message)
    root.destroy()


# Google DLP integration

def detect_pii_with_dlp(text):
    client = dlp.DlpServiceClient()
    parent = f"projects/YOUR_PROJECT_ID"

    item = {"value": text}
    config = {
        "info_types": [{"name": "ALL_BASIC"}],
        "include_quote": True,
        "min_likelihood": "POSSIBLE"
    }

    response = client.inspect_content(request={"parent": parent, "item": item, "inspect_config": config})

    findings = []
    for result in response.result.findings:
        findings.append((result.info_type.name, result.quote))
    return findings


# Scanning function

def scan_folder(queue, folder_path):
    df = pd.DataFrame(columns=["Timestamp", "PII Type", "Masked PII", "SHA256 Hash", "File Name", "Line Number"])

    for fname in os.listdir(folder_path):
        full = os.path.join(folder_path, fname)
        if not os.path.isfile(full):
            continue

        text = extract_text_from_any_file(full)
        if queue:
            queue.put({"type": "file_scanned", "file_name": fname})

        if not text.strip():
            print(f"Skipping (no text): {fname}")
            continue

        print(f"\n--- Scanning {fname} ---")
        detected = False

        # Local regex detection
        for pii, pat in patterns.items():
            for idx, m in enumerate(re.findall(pat, text, flags=re.IGNORECASE), start=1):
                detected = True
                print(f"  ✅ {pii}: {m}")
                df = log_detected_pii(df, pii, m, fname, idx)

        # Google DLP detection
        for idx, (pii_type, quote) in enumerate(detect_pii_with_dlp(text), start=1):
            print(f"  🔍 [DLP] {pii_type}: {quote}")
            df = log_detected_pii(df, pii_type, quote, fname, idx)
            detected = True

        if not detected:
            print(f"  ⚪ No PII in {fname}")

    if not df.empty:
        summary = df.to_string(index=False)
        show_popup("Scan Complete", "PII found—sending email alert now…")
        send_email_alert("✅ PII Detection Alert", f"Dear Team,\n\n{summary}\n\nRegards", df)
    else:
        print("✅ No PII found.")


# Monitor
class PIIFileHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory:
            return
        path = event.src_path
        print("Detected new file:", path)
        text = try_extract_text_with_retries(path)
        if not text.strip():
            print("  → No text extracted.")
            return

        if any(re.search(p, text, flags=re.IGNORECASE) for p in patterns.values()) or detect_pii_with_dlp(text):
            show_popup("PII Detected", f"'{os.path.basename(path)}' contains PII and cannot stay here!")
            try:
                os.remove(path)
                print("  → Deleted:", path)
            except Exception as e:
                print("  → Delete failed:", e)


def try_extract_text_with_retries(path, retries=5, delay=1):
    for _ in range(retries):
        text = extract_text_from_any_file(path)
        if text.strip():
            return text
        time.sleep(delay)
    return ""


def start_monitor(folder):
    observer = Observer()
    observer.schedule(PIIFileHandler(), folder, recursive=False)
    observer.start()
    print("Monitoring folder:", folder)
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


# Start monitoring
monitor_thread = threading.Thread(target=start_monitor, args=(USB_MONITOR_PATH,), daemon=True)
monitor_thread.start()

if __name__ == "__main__":
    scan_folder(None, FOLDER_PATH)
    print("Monitoring in background. Ctrl+C to exit.")
    try:
        while True:
            time.sleep(10)
    except KeyboardInterrupt:
        print("Exiting.")
